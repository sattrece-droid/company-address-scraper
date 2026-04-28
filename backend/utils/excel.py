"""
Excel I/O utilities for parsing input files and generating output files.
Handles optional columns (zip code, website) and generates structured output.
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import List, Dict, Any, Optional
from pathlib import Path
import pandas as pd


class ExcelHandler:
    """Handle Excel file parsing and generation."""
    
    @staticmethod
    def parse_input_file(file_path: str) -> List[Dict[str, Optional[str]]]:
        """
        Parse input Excel file and extract company information.
        
        Expected columns:
        - Company Name (required)
        - Zip Code (optional)
        - Website (optional)
        
        Returns list of dicts with keys: company_name, zip_code, website
        """
        try:
            # Use pandas for flexible column detection
            df = pd.read_excel(file_path)
            
            # Normalize column names (case-insensitive, strip whitespace)
            df.columns = df.columns.str.strip().str.lower()
            
            # Map various column name variations
            column_mapping = {
                'company_name': ['company name', 'company', 'name', 'business name'],
                'zip_code': ['zip code', 'zip', 'zipcode', 'postal code', 'postalcode'],
                'website': ['website', 'web', 'url', 'site']
            }
            
            # Find actual column names
            actual_columns = {}
            for target, variations in column_mapping.items():
                for col in df.columns:
                    if col in variations:
                        actual_columns[target] = col
                        break
            
            # Ensure company_name column exists
            if 'company_name' not in actual_columns:
                raise ValueError(
                    "Input Excel must have a 'Company Name' column. "
                    f"Found columns: {', '.join(df.columns)}"
                )
            
            # Extract data
            companies = []
            for _, row in df.iterrows():
                company_name = row[actual_columns['company_name']]
                
                # Skip empty rows
                if pd.isna(company_name) or str(company_name).strip() == '':
                    continue
                
                companies.append({
                    'company_name': str(company_name).strip(),
                    'zip_code': str(row[actual_columns['zip_code']]).strip() 
                                if 'zip_code' in actual_columns and not pd.isna(row[actual_columns['zip_code']])
                                else None,
                    'website': str(row[actual_columns['website']]).strip()
                              if 'website' in actual_columns and not pd.isna(row[actual_columns['website']])
                              else None
                })
            
            return companies
        
        except Exception as e:
            raise ValueError(f"Error parsing Excel file: {str(e)}")
    
    @staticmethod
    def generate_output_file(
        companies_data: List[Dict[str, Any]],
        output_path: str
    ) -> str:
        """
        Generate output Excel file with structured address data.
        
        Output columns:
        - Company Name
        - Input Zip Code (if provided)
        - Status
        - Match Confidence
        - Location Name
        - Street Address
        - City
        - State/Province
        - Zip Code
        - Country
        - Cached (Yes/No)
        
        Args:
            companies_data: List of company results with addresses
            output_path: Path to save the output Excel file
        
        Returns:
            Path to the generated file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Company Addresses"
        
        # Define headers
        headers = [
            "Company Name",
            "Input Zip Code",
            "Status",
            "Match Confidence",
            "Location Name",
            "Street Address",
            "City",
            "State/Province",
            "Zip Code",
            "Country",
            "Cached"
        ]
        
        # Style headers
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        # Write data
        current_row = 2
        for company in companies_data:
            company_name = company.get('company_name', '')
            input_zip = company.get('input_zip_code', '')
            status = company.get('status', 'unknown')
            confidence = company.get('confidence', 'N/A')
            cached = "Yes" if company.get('cached', False) else "No"
            
            addresses = company.get('addresses', [])
            
            if addresses:
                # Write one row per location
                for address in addresses:
                    ws.cell(row=current_row, column=1, value=company_name)
                    ws.cell(row=current_row, column=2, value=input_zip)
                    ws.cell(row=current_row, column=3, value=status)
                    ws.cell(row=current_row, column=4, value=confidence)
                    ws.cell(row=current_row, column=5, value=address.get('name', ''))
                    ws.cell(row=current_row, column=6, value=address.get('address', ''))
                    ws.cell(row=current_row, column=7, value=address.get('city', ''))
                    ws.cell(row=current_row, column=8, value=address.get('state', ''))
                    ws.cell(row=current_row, column=9, value=address.get('zip', ''))
                    ws.cell(row=current_row, column=10, value=address.get('country', ''))
                    ws.cell(row=current_row, column=11, value=cached)
                    current_row += 1
            else:
                # No addresses found - write company info only
                ws.cell(row=current_row, column=1, value=company_name)
                ws.cell(row=current_row, column=2, value=input_zip)
                ws.cell(row=current_row, column=3, value=status)
                ws.cell(row=current_row, column=4, value=confidence)
                ws.cell(row=current_row, column=11, value=cached)
                current_row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Ensure output directory exists
        Path(output_path).parent.mkdir(parents=True, exist_ok=True)
        
        # Save workbook
        wb.save(output_path)
        return output_path
    
    @staticmethod
    def validate_input_file(file_path: str) -> Dict[str, Any]:
        """
        Validate input Excel file structure and content.
        
        Returns validation result with status and details.
        """
        try:
            companies = ExcelHandler.parse_input_file(file_path)
            
            if not companies:
                return {
                    "valid": False,
                    "error": "No valid company entries found in Excel file"
                }
            
            return {
                "valid": True,
                "company_count": len(companies),
                "has_zip_codes": any(c.get('zip_code') for c in companies),
                "has_websites": any(c.get('website') for c in companies)
            }
        
        except Exception as e:
            return {
                "valid": False,
                "error": str(e)
            }